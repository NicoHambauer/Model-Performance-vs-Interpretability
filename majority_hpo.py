import openpyxl
from collections import defaultdict, Counter

import pandas as pd
import warnings

# This is a script that parses the results of the HPO runs and finds the best config for each model and dataset
# Required is an Excel file with the following structure:
# - Each sheet contains the results of one HPO run
# - One has to specify the column and row indices of the cells of interest (see below)
# - The specified cells are used to find the best config for each model and dataset
# - The best config is defined as the config that occurs most often in the specified cells
# - The best config is stored in a dataframe and exported as an Excel file


# Helper function to remove \n inside square brackets as happening in MLP parameters
def remove_newlines_inside_square_brackets(input_string):
    inside_brackets = False
    start_index = None
    part_to_be_replaced = []
    replacing_part = []
    for i, char in enumerate(input_string):
        if char == '[' and not inside_brackets:
            inside_brackets = True
            start_index = i
        elif char == ']' and inside_brackets:
            inside_brackets = False
            part_to_be_replaced.append(input_string[start_index:i+1])
            replacing_part.append(input_string[start_index:i+1].replace('\n', ' '))
    result = input_string
    for index, old_part in enumerate(part_to_be_replaced):
        result = result.replace(old_part, replacing_part[index])
    return result

# Open the Excel file
workbook = openpyxl.load_workbook('hpo_best_config_all_folds_reruns.xlsx')

# Define column and row indices to look into
# take care, that model_names ordering an row_names ordering is equal to ordering in excel file
model_names = ['LR', 'DT', 'RF', 'XGB', 'MLP', 'PYGAM', 'EBM', 'NAM', 'GAMINET', 'EXNN',  'IGANN', 'ELASTICNET']
row_names = ['weather', 'stroke', 'adult', 'telco', 'college', 'fico', 'bank', 'airline', 'compas', 'water', 'car', 'crab', 'diamond', 'medical', 'productivity', 'student', 'crimes', 'bike', 'housing', 'wine']
col_indices = range(2, 14)
row_indices = range(3, 23)

# create empty dataframe to store the best configs
best_configs = pd.DataFrame(index=row_names, columns=model_names)
# Iterate over each cell of interest
for row_index in row_indices:
    for col_index in col_indices:

        # Define dictionaries to store parsed data
        dict_list = []
        for i in range(5):
            dict_list.append(defaultdict(int))

        # Iterate over each sheet in the workbook
        for sheet_index, sheet_name in enumerate(workbook.sheetnames):
            sheet = workbook[sheet_name]
            cell_value = sheet.cell(row=row_index, column=col_index).value
            if cell_value:
                # Parse the string into a dictionary
                try:
                    # If string container squared bracked, check if \n are inside of it and remove them
                    if "[" in cell_value:
                       cell_value = remove_newlines_inside_square_brackets(cell_value)
                    parsed_dict = dict(line.split(': ') for line in cell_value.split('\n'))
                except ValueError:
                    continue
                # Update the corresponding dictionary in dict_list
                for key in parsed_dict:
                    # For whatever reason the chars '_x000D_' are added to some values, remove them
                    # I guess it is introducing a newline in excel (but also in the LaTeX table)
                    # we remove it here to be able to cast the values to their correct type int or float
                    # later, we add it again to the string to have line breaks in the LaTeX table
                    if '_x000D_' in parsed_dict[key]:
                        parsed_dict[key] = parsed_dict[key].replace('_x000D_', '')

                    # Try to encode the value as int or float, if this fails, use string
                    try:
                        value = int(parsed_dict[key])
                    except ValueError:
                        try:
                            value = float(parsed_dict[key])
                        except ValueError:
                            value = parsed_dict[key].strip("' ")
                    key = key.strip("' ")
                    dict_list[sheet_index][key] = value

        # Find the value with the most occurrences for each key by iterating over the parameter keys in dict_list
        final_dict = {}
        for key in dict_list[0]:
            value_list = []
            # for each parameter key, append the values of the HPO runs to value_list
            for run in range(len(dict_list)):
                value_list.append(dict_list[run][key])
            # Find the value with the most occurrences
            top_1_value = Counter(value_list).most_common(1)
            # If there is a tie, print a warning
            if len(Counter(value_list).most_common()) > 1 and Counter(value_list).most_common()[0][1] == Counter(value_list).most_common()[1][1]:
                warnings.warn(f'There is a tie in the best config for the following model and dataset: \n \
                Model: {model_names[col_indices.index(col_index)]} Dataset: {row_names[row_indices.index(row_index)]}\n \
                {Counter(value_list).most_common()}')
                key = key + '*'
            final_dict[key] = top_1_value[0][0]
        # Print the final dictionary
        print(final_dict, f'model: {model_names[col_indices.index(col_index)]}, dataset: {row_names[row_indices.index(row_index)]}')

        # here we add _x000D_ again to the string to have line breaks in the LaTeX table, as described above
        cleaned_cell_data = str(final_dict).strip('{}').replace(",", ",_x000D_")
        best_configs[model_names[col_indices.index(col_index)]][row_names[row_indices.index(row_index)]] = cleaned_cell_data
        final_dict = {}

# sort columns and rows in a specifc way - as needed for the GAM Compare paper
best_configs = best_configs[['PYGAM', 'EBM', 'NAM', 'GAMINET', 'EXNN', 'IGANN', 'LR', 'ELASTICNET', 'DT', 'RF', 'XGB', 'MLP']]
index_order = ['college', 'water', 'stroke', 'telco', 'compas', 'fico', 'adult', 'bank', 'airline', 'weather', 'car', 'student', 'productivity', 'medical', 'crimes',  'crab', 'wine', 'bike', 'housing', 'diamond']
best_configs = best_configs.reindex(index_order)
best_configs.to_excel('best_configs.xlsx')
best_configs.to_latex('best_configs.tex')
best_configs.to_csv('best_configs.csv')




